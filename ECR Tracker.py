# Michael Tillapaugh
# Updated September 13, 2017

import os
import openpyxl

# sheet.cell(row=1, column=2).value is the same as saying sheet['B1'].value
# to get the value of a specific cell


def ECR_tracker_input(n, filename):
    os.chdir("P:\\E-Architecture\\Electrical 12V Systems\\12V Harness\\2D Drawings\\FF91 Gamma\\ECR Files\\ECRs")
    wb = openpyxl.load_workbook(filename, data_only=True)
    sheet = wb.get_sheet_by_name('pg 1')
    cell_list = ['AC1', 'AF2', 'AT1', 'AU2', 'B11', 'H3', 'A76', 'AP3', 'BV17']
    #  Update this list to change which cells are being copied from an ECR
    Harness = sheet[cell_list[0]].value
    Implementation = sheet[cell_list[1]].value
    ECR_No = sheet[cell_list[2]].value
    JIRA_No = sheet[cell_list[3]].value
    Description = sheet[cell_list[4]].value
    Priority = sheet[cell_list[5]].value
    Reason_For_Change = sheet[cell_list[6]].value
    Expected_Implementation_Timing = sheet[cell_list[7]].value
    Status = sheet[cell_list[8]].value
    # Need one variable above for each cell in cell_list

    os.chdir("P:\\E-Architecture\\Electrical 12V Systems\\12V Harness\\2D Drawings\\FF91 Gamma\\ECR Files")
    wb = openpyxl.load_workbook('ECR Tracker Master.xlsx')
    sheet = wb.get_sheet_by_name('Sheet')
    sheet.cell(row=n, column=1).value = filename[:-5]
    sheet.cell(row=n, column=2).value = Harness
    sheet.cell(row=n, column=3).value = Implementation
    sheet.cell(row=n, column=4).value = ECR_No
    sheet.cell(row=n, column=5).value = JIRA_No
    sheet.cell(row=n, column=6).value = Description
    sheet.cell(row=n, column=7).value = Priority
    sheet.cell(row=n, column=8).value = Reason_For_Change
    sheet.cell(row=n, column=9).value = Expected_Implementation_Timing
    sheet.cell(row=n, column=10).value = Status
    # The rows above 'paste' each variable into a specific cell in the new file

    wb.save('ECR Tracker Master.xlsx')


os.chdir("P:\\E-Architecture\\Electrical 12V Systems\\12V Harness\\2D Drawings\\FF91 Gamma\\ECR Files\\ECRs")

list = []
dirListing = os.listdir()
# sets 'dirListing' to an array of all files in the current working directory

for item in dirListing:
    if ".xlsx" in item:
        list.append(item)
# Goes through 'dirListing' and adds anything that has .xlsx to 'list'

count = 2
for i in list:
    ECR_tracker_input(count, i)
    count += 1
# For each file in 'list' go through the ECR_tracker_input function
